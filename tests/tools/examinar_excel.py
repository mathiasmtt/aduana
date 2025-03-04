#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
Script para examinar la estructura de un archivo Excel de aranceles.

Uso:
    python examinar_excel.py archivo.xlsx

Salida:
    Muestra información sobre las hojas y columnas del archivo Excel.
"""

import os
import sys
import pandas as pd
from openpyxl import load_workbook

def examinar_excel(archivo):
    """
    Examina la estructura de un archivo Excel y muestra información relevante
    
    Args:
        archivo (str): Ruta al archivo Excel
    """
    print(f"\nExaminando archivo: {archivo}")
    try:
        # Cargar el libro de trabajo para examinar las hojas
        workbook = load_workbook(archivo, read_only=True)
        
        # Mostrar las hojas disponibles
        print(f"\nHojas disponibles: {workbook.sheetnames}")
        
        # Examinar cada hoja
        for sheet_name in workbook.sheetnames:
            print(f"\n{'='*40}")
            print(f"Hoja: {sheet_name}")
            print(f"{'='*40}")
            
            # Cargar la hoja como DataFrame
            df = pd.read_excel(archivo, sheet_name=sheet_name, nrows=10)
            
            # Mostrar las primeras filas para entender la estructura
            print("\nPrimeras filas:")
            print(df.head(5))
            
            # Mostrar información sobre las columnas
            print("\nInformación de columnas:")
            for col in df.columns:
                print(f"  - {col} (Tipo: {df[col].dtype})")
            
            # Buscar patrones específicos en las columnas
            print("\nBuscando columnas relevantes:")
            for col in df.columns:
                col_str = str(col).upper()
                if 'NCM' in col_str:
                    print(f"  - Posible columna NCM: {col}")
                elif 'AEC' in col_str or 'ARANCEL' in col_str or 'EXTERNO' in col_str:
                    print(f"  - Posible columna AEC: {col}")
                elif 'DESCRIP' in col_str:
                    print(f"  - Posible columna de descripción: {col}")
    
    except Exception as e:
        print(f"Error al examinar el archivo: {e}")

def main():
    if len(sys.argv) < 2:
        print("Uso: python examinar_excel.py archivo.xlsx")
        return 1
    
    archivo = sys.argv[1]
    if not os.path.exists(archivo):
        print(f"Error: El archivo {archivo} no existe.")
        return 1
    
    examinar_excel(archivo)
    return 0

if __name__ == "__main__":
    sys.exit(main())
