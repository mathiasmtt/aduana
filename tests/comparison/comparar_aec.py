#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
Script para comparar los Aranceles Externos Comunes (AEC) entre dos archivos Excel
y encontrar diferencias.

Uso:
    python comparar_aec.py archivo1.xlsx archivo2.xlsx

Salida:
    Muestra los NCMs donde el AEC ha cambiado, con los valores antes y después.
"""

import os
import sys
import pandas as pd
import argparse
from openpyxl import load_workbook

def cargar_excel(archivo, sheet_name=None):
    """
    Carga un archivo Excel de aranceles y devuelve un DataFrame con los datos relevantes
    
    Args:
        archivo (str): Ruta al archivo Excel
        sheet_name (str): Nombre de la hoja (opcional)
        
    Returns:
        DataFrame con los datos
    """
    try:
        # Primero intentamos determinar el nombre de la hoja si no se proporcionó
        if not sheet_name:
            workbook = load_workbook(archivo, read_only=True)
            sheet_name = workbook.sheetnames[0]  # Usar la primera hoja por defecto
        
        # Usamos pandas para cargar el Excel
        return pd.read_excel(archivo, sheet_name=sheet_name)
    except Exception as e:
        print(f"Error al cargar el archivo {archivo}: {e}")
        return None

def limpiar_datos(df):
    """
    Limpia y prepara los datos del DataFrame para la comparación
    
    Args:
        df (DataFrame): DataFrame con los datos del Excel
        
    Returns:
        DataFrame limpio y preparado
    """
    # Verificar que las columnas necesarias existan
    col_ncm = None
    col_aec = None
    
    # Buscar las columnas que contienen NCM y AEC
    for col in df.columns:
        col_str = str(col).upper()
        if 'NCM' in col_str:
            col_ncm = col
        elif 'AEC' in col_str or 'ARANCEL EXTERNO' in col_str:
            col_aec = col
    
    if not col_ncm or not col_aec:
        print("No se encontraron las columnas necesarias (NCM y AEC)")
        return None
    
    # Crear un nuevo DataFrame con solo las columnas necesarias
    df_clean = df[[col_ncm, col_aec]].copy()
    
    # Renombrar columnas para estandarizar
    df_clean.columns = ['NCM', 'AEC']
    
    # Eliminar filas con NCM vacío o no válido
    df_clean = df_clean.dropna(subset=['NCM'])
    
    # Convertir NCM a string y limpiar
    df_clean['NCM'] = df_clean['NCM'].astype(str).str.strip()
    
    # Convertir AEC a numérico
    df_clean['AEC'] = pd.to_numeric(df_clean['AEC'], errors='coerce')
    
    # Usar NCM como índice para facilitar la comparación
    return df_clean.set_index('NCM')

def comparar_aec(df1, df2):
    """
    Compara los AEC entre dos DataFrames y encuentra las diferencias
    
    Args:
        df1 (DataFrame): DataFrame del primer archivo
        df2 (DataFrame): DataFrame del segundo archivo
        
    Returns:
        DataFrame con las diferencias
    """
    # Verificar que ambos DataFrames existen
    if df1 is None or df2 is None:
        return None
    
    # Encontrar NCMs comunes
    ncm_comunes = set(df1.index) & set(df2.index)
    
    # Crear un DataFrame vacío para almacenar las diferencias
    diferencias = []
    
    # Comparar AEC para cada NCM común
    for ncm in ncm_comunes:
        aec1 = df1.loc[ncm, 'AEC']
        aec2 = df2.loc[ncm, 'AEC']
        
        # Si los AEC son diferentes, agregar a las diferencias
        if aec1 != aec2:
            diferencias.append({
                'NCM': ncm,
                'AEC_anterior': aec1,
                'AEC_nuevo': aec2,
                'Diferencia': aec2 - aec1
            })
    
    # Convertir a DataFrame y ordenar por la magnitud de la diferencia
    df_diferencias = pd.DataFrame(diferencias)
    if not df_diferencias.empty:
        df_diferencias = df_diferencias.sort_values(by='Diferencia', ascending=False)
    
    return df_diferencias

def main():
    parser = argparse.ArgumentParser(description='Comparar AEC entre dos archivos Excel')
    parser.add_argument('--archivo1', required=True, help='Ruta al primer archivo Excel (versión anterior)')
    parser.add_argument('--archivo2', required=True, help='Ruta al segundo archivo Excel (versión nueva)')
    parser.add_argument('--hoja1', help='Nombre de la hoja en el primer archivo (opcional)')
    parser.add_argument('--hoja2', help='Nombre de la hoja en el segundo archivo (opcional)')
    parser.add_argument('--ejemplos', type=int, default=10, help='Número de ejemplos a mostrar (default: 10)')
    
    args = parser.parse_args()
    
    # Cargar los datos de ambos archivos
    print(f"Cargando datos del archivo anterior: {args.archivo1}")
    df1 = cargar_excel(args.archivo1, args.hoja1)
    if df1 is not None:
        df1 = limpiar_datos(df1)
        print(f"  - {len(df1)} registros encontrados")
    
    print(f"Cargando datos del archivo nuevo: {args.archivo2}")
    df2 = cargar_excel(args.archivo2, args.hoja2)
    if df2 is not None:
        df2 = limpiar_datos(df2)
        print(f"  - {len(df2)} registros encontrados")
    
    # Comparar los AEC
    print("Comparando AEC...")
    diferencias = comparar_aec(df1, df2)
    
    if diferencias is not None and not diferencias.empty:
        print(f"\nSe encontraron {len(diferencias)} NCMs con cambios en el AEC:")
        
        # Mostrar los ejemplos solicitados
        ejemplos = min(args.ejemplos, len(diferencias))
        print(f"\nMostrando {ejemplos} ejemplos de NCMs con mayores cambios:")
        
        pd.set_option('display.max_rows', None)
        pd.set_option('display.max_columns', None)
        pd.set_option('display.width', 120)
        pd.set_option('display.precision', 2)
        
        print(diferencias.head(ejemplos).to_string(index=False))
        
        return 0
    else:
        print("No se encontraron diferencias en el AEC entre los archivos.")
        return 1

if __name__ == "__main__":
    sys.exit(main())
